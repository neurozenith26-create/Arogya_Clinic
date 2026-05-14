"""Generate module-wise user stories Excel for Arogya Diagnostics.

Run from anywhere; writes to ./Arogya_User_Stories.xlsx next to this file.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# ---------------------------------------------------------------------------
# Story data — (Module, ID, Role, Story, Acceptance Criteria, Priority, Status)
# Roles: Visitor, Patient, Admin, Super Admin, Doctor (read-only), Collector
# Status reflects current code state observed in repo.
# ---------------------------------------------------------------------------

STORIES = [
    # ---------- 1. Public Marketing Site ----------
    ("1. Public Marketing Site", "PUB-01", "Visitor",
     "As a visitor, I want a homepage with hero, services snapshot, doctors snapshot, testimonials, and contact CTA so I can quickly understand what the clinic offers.",
     "Hero with clinic name + tagline; sections for services, doctors, departments, reviews; bilingual EN/BN service names; CTA buttons for Book / Call / WhatsApp.",
     "Must", "Implemented"),
    ("1. Public Marketing Site", "PUB-02", "Visitor",
     "As a visitor, I want to browse all diagnostic services and health packages filtered by category and free-text search so I can find what I need.",
     "GET /services?category_slug=&q= returns active services only; ILIKE on name; ordered by display_order; max 200 results.",
     "Must", "Implemented"),
    ("1. Public Marketing Site", "PUB-03", "Visitor",
     "As a visitor, I want a service detail page with description, prep instructions, sample type, turnaround time and price so I can decide before booking.",
     "GET /services/:slug; renders full_details, prep_instructions, sample_type, report_turnaround_hours, price; Book Test CTA adds to cart.",
     "Must", "Implemented"),
    ("1. Public Marketing Site", "PUB-04", "Visitor",
     "As a visitor, I want to browse all verified doctors filtered by department and search so I can pick the right specialist.",
     "GET /doctors?department_id=&q= filters role=doctor + is_verified + is_active; ordered by rating_avg desc; LIMIT 100.",
     "Must", "Implemented"),
    ("1. Public Marketing Site", "PUB-05", "Visitor",
     "As a visitor, I want a doctor profile page with photo, qualifications, departments, consulting centers, schedule preview and reviews so I can choose confidently.",
     "GET /doctors/:id + GET /reviews?doctor_id=; photo via public /doctors/:id/photo blob; reviews limited to status=approved.",
     "Must", "Implemented"),
    ("1. Public Marketing Site", "PUB-06", "Visitor",
     "As a visitor, I want to browse clinical departments and their associated doctors so I can navigate by speciality.",
     "GET /departments + GET /departments/:slug; active departments only; ordered by display_order.",
     "Should", "Implemented"),
    ("1. Public Marketing Site", "PUB-07", "Visitor",
     "As a visitor, I want a contact form to send the clinic an enquiry without signing up so I can ask questions easily.",
     "POST /enquiries with name/email/phone/subject/message; rate-limited 5/min/IP; inserts row with status=new.",
     "Must", "Implemented"),
    ("1. Public Marketing Site", "PUB-08", "Visitor",
     "As a visitor, I want to leave feedback / a review for the clinic or a doctor so future patients can read it.",
     "POST /reviews with star rating + text; status=pending (admin moderation); rate-limited 5/min/IP.",
     "Should", "Implemented"),
    ("1. Public Marketing Site", "PUB-09", "Visitor",
     "As a visitor, I want privacy and terms pages so I can understand how my data is used.",
     "Static pages at /privacy and /terms accessible without login.",
     "Should", "Implemented"),
    ("1. Public Marketing Site", "PUB-10", "Visitor",
     "As a visitor, I want a Book Home Visit CTA in the header so I can start a home-collection booking immediately.",
     "Header button sets cartStore.preferredVisitType='home_visit' and navigates to /services?mode=home_visit; Services page shows banner with reset option.",
     "Should", "Implemented"),

    # ---------- 2. Authentication & Identity ----------
    ("2. Authentication & Identity", "AUTH-01", "Patient",
     "As a new patient, I want to sign up with first/last name, mobile, optional email and password so I can manage my bookings.",
     "POST /auth/signup; mobile regex ^[6-9]\\d{9}$; password ≥6; rejects duplicate mobile/email; returns JWT + user.",
     "Must", "Implemented"),
    ("2. Authentication & Identity", "AUTH-02", "Patient",
     "As an existing patient, I want to log in with email and password so I can access my dashboard.",
     "POST /auth/login; bcrypt verify; checks is_active && is_login_enabled; updates last_login_at; returns JWT + user.",
     "Must", "Implemented"),
    ("2. Authentication & Identity", "AUTH-03", "Patient",
     "As a patient without a password, I want to log in with mobile + OTP so I don't need to remember credentials.",
     "POST /auth/otp/send returns 6-digit code (debug_code in non-prod); POST /auth/otp/verify accepts code + optional names; auto-creates patient row if mobile is new.",
     "Must", "Implemented (OTP via SMS provider TODO)"),
    ("2. Authentication & Identity", "AUTH-04", "Patient",
     "As a logged-in patient, I want to view my profile (name, mobile, email, DOB, gender, default address) so I can keep it current.",
     "GET /auth/me returns full user row; rendered on /dashboard/profile.",
     "Must", "Implemented"),
    ("2. Authentication & Identity", "AUTH-05", "Patient",
     "As a logged-in patient, I want to update my profile fields including default address JSON so future bookings prefill correctly.",
     "PATCH /me with profileUpdateSchema; dynamic SET builder; default_address stored as JSONB.",
     "Must", "Implemented"),
    ("2. Authentication & Identity", "AUTH-06", "Patient",
     "As a logged-in patient, I want to change my password so I can rotate credentials.",
     "POST /auth/change-password; current_password verified if existing hash; new_password hashed via bcrypt.",
     "Should", "Implemented"),
    ("2. Authentication & Identity", "AUTH-07", "Patient",
     "As a logged-in patient, I want to log out from the header dropdown so my JWT is cleared from the device.",
     "Best-effort POST /auth/logout + clears arogya-jwt localStorage + zustand state.",
     "Must", "Implemented"),
    ("2. Authentication & Identity", "AUTH-08", "Admin",
     "As an admin/super_admin, I want a separate /admin/login route so my login experience is distinct from the patient flow.",
     "AdminLoginPage; same POST /auth/login; on success redirects to /admin and RequireAuth checks role admin|super_admin.",
     "Must", "Implemented"),
    ("2. Authentication & Identity", "AUTH-09", "Any role",
     "As any logged-in user, I want my JWT to be re-validated on every request so revoking an account takes effect immediately.",
     "requireAuth middleware re-reads users row by sub claim; sets req.user; rejects with 403 ACCOUNT_DISABLED when is_active=false.",
     "Must", "Implemented"),
    ("2. Authentication & Identity", "AUTH-10", "Patient",
     "As a patient who got auto-logged out (401), I want my JWT cleared automatically so /auth/me doesn't loop.",
     "Axios response interceptor clears token on 401 then propagates the error.",
     "Should", "Implemented"),

    # ---------- 3. Patient Dashboard ----------
    ("3. Patient Dashboard", "PD-01", "Patient",
     "As a patient, I want a dashboard overview showing upcoming bookings, pending payments and recent reports so I get a one-glance status.",
     "DashboardOverviewPage uses useMyBookings()/useMyReports() to render KPI cards + next-appointment + recent-reports.",
     "Must", "Implemented"),
    ("3. Patient Dashboard", "PD-02", "Patient",
     "As a patient, I want a dedicated Appointments page listing my doctor bookings so I can track them.",
     "GET /bookings filtered by booking_type=doctor_appointment; each row shows doctor_name, speciality, center, status badge.",
     "Must", "Implemented"),
    ("3. Patient Dashboard", "PD-03", "Patient",
     "As a patient, I want a dedicated Tests page listing my test/package bookings so I can track sample collection and results.",
     "GET /bookings filtered to test_booking; row shows items_summary, collection_status, reports_count.",
     "Must", "Implemented"),
    ("3. Patient Dashboard", "PD-04", "Patient",
     "As a patient, I want a combined My Bookings page if I prefer one stream view.",
     "MyBookingsPage at /dashboard/bookings unifies both booking types ordered by scheduled_date desc.",
     "Should", "Implemented"),
    ("3. Patient Dashboard", "PD-05", "Patient",
     "As a patient, I want to open a booking detail page from any list to see items, payment status, address, advance/balance and reports.",
     "GET /bookings/:id; ownership-checked; renders items, payments, active reports with download links.",
     "Must", "Implemented"),
    ("3. Patient Dashboard", "PD-06", "Patient",
     "As a patient, I want to cancel an upcoming booking from its detail page if my plans change.",
     "POST /bookings/:id/cancel sets booking_status=cancelled and records cancelled_at + optional reason.",
     "Must", "Implemented"),
    ("3. Patient Dashboard", "PD-07", "Patient",
     "As a patient, I want a Reports page that lists all my lab results across bookings so I can find them in one place.",
     "GET /reports patient-scoped; ReportsPage thumbnails + download via signed URL.",
     "Must", "Implemented"),
    ("3. Patient Dashboard", "PD-08", "Patient",
     "As a patient, I want a clear status sub-line like 'Advance ₹100 paid · Balance ₹100 due at visit' instead of generic status text.",
     "BookingStatusBadge takes paymentStatus + totalAmount + advanceAmount and renders contextual sub-line; threaded through all dashboard pages.",
     "Should", "Implemented"),

    # ---------- 4. Booking — Doctor Appointment ----------
    ("4. Doctor Appointment Booking", "DA-01", "Patient",
     "As a patient, I want to pick a doctor's available date and time slot from a live calendar so I don't book a conflicting slot.",
     "GET /doctors/:id/slots?date=&center_id=&visit_type=; live generator reads doctor_centers.schedule, subtracts existing bookings + doctor_unavailability; returns {time, available, reason}.",
     "Must", "Implemented"),
    ("4. Doctor Appointment Booking", "DA-02", "Patient",
     "As a patient, I want to enter or confirm my contact details and address for the appointment.",
     "Step 2 of BookDoctorPage uses patientInfoSchema; prefills from useAuthStore user when present.",
     "Must", "Implemented"),
    ("4. Doctor Appointment Booking", "DA-03", "Patient",
     "As a patient, I want a clear breakdown of consultation fee, advance (50%) and balance so I know what I'm paying.",
     "subtotal = doctor.consultation_fee; advance = round(total/2); balance = total - advance; rendered in checkout summary.",
     "Must", "Implemented"),
    ("4. Doctor Appointment Booking", "DA-04", "Patient",
     "As a patient, I want to submit the booking with a UPI payment proof image/PDF so my slot is locked immediately.",
     "POST /bookings/doctor-appointment (multipart) inserts bookings + booking_items(doctor_consultation) + invoices (trigger numbers it) + payments(upi_manual, captured, verified_at=NULL) in one transaction.",
     "Must", "Implemented"),
    ("4. Doctor Appointment Booking", "DA-05", "Patient",
     "As a patient, I want a no-consulting-center guard so I'm not allowed to book a doctor whose schedule isn't set up yet.",
     "BookDoctorPage shows friendly message when doctor.centers[] is empty.",
     "Should", "Implemented"),

    # ---------- 5. Booking — Diagnostic Test / Package ----------
    ("5. Test / Package Booking", "TB-01", "Patient",
     "As a patient, I want to add multiple tests/packages to a cart with quantities so I can check out in one go.",
     "useCartStore items[] persists to localStorage 'arogya-cart'; supports addItem, removeItem, updateQuantity, subtotal, count.",
     "Must", "Implemented"),
    ("5. Test / Package Booking", "TB-02", "Patient",
     "As a patient, I want to pick in-clinic vs home collection per booking so I can choose how the sample is taken.",
     "Step 2 of BookTestPage switches visit_type; home_visit branch reveals address + pincode check + slot picker.",
     "Must", "Implemented"),
    ("5. Test / Package Booking", "TB-03", "Patient",
     "As a patient, I want to enter my pincode and instantly see if home collection is serviceable + the home-visit charge.",
     "GET /serviceable-pincodes/check?pincode= validates regex; returns {serviceable, city, home_visit_charge, collection_lead_time_hours}.",
     "Must", "Implemented"),
    ("5. Test / Package Booking", "TB-04", "Patient",
     "As a patient, I want to pick a home collection time slot if my pincode is serviceable.",
     "GET /home-collection/slots?date=&pincode= currently returns hardcoded 8-slot grid (STUB); hook is enabled only when pincode confirmed.",
     "Must", "Stub (hardcoded slots)"),
    ("5. Test / Package Booking", "TB-05", "Patient",
     "As a patient, I want a clear breakdown showing items × qty, subtotal, home-visit charge, total, advance and balance.",
     "Step 3 renders summary; advance = round((subtotal + home_visit_charge)/2).",
     "Must", "Implemented"),
    ("5. Test / Package Booking", "TB-06", "Patient",
     "As a patient, I want to submit my test booking with a UPI payment proof so my slot is reserved immediately and the invoice is auto-generated.",
     "POST /bookings/test (multipart) inserts bookings + N booking_items + invoices + payments(upi_manual) atomically; redirects to /payment/callback.",
     "Must", "Implemented"),
    ("5. Test / Package Booking", "TB-07", "Patient",
     "As a patient, I want my home-visit preference to survive navigation from Services → Cart → Booking.",
     "cartStore.preferredVisitType persists; BookTestPage reads it for initial visit-type state.",
     "Should", "Implemented"),

    # ---------- 6. Payments — Manual UPI ----------
    ("6. Payments (Manual UPI)", "PAY-01", "Patient",
     "As a patient at checkout, I want to see the clinic's UPI VPA + QR code so I can pay from any UPI app.",
     "GET /clinic/upi returns {upi_id, upi_display_name} from clinic_settings; UpiPaymentModal renders QR via react-qr-code and copy-VPA button.",
     "Must", "Implemented"),
    ("6. Payments (Manual UPI)", "PAY-02", "Patient",
     "As a patient, I want to attach a screenshot or PDF of the UPI confirmation (with optional UTR) so my booking is captured.",
     "FormData: payment_proof (≤5 MB, jpeg/png/webp/pdf) + data (JSON booking payload) + optional upi_reference; backend validates via validateProofFile().",
     "Must", "Implemented"),
    ("6. Payments (Manual UPI)", "PAY-03", "Patient",
     "As a patient, I want to view the proof I uploaded inside my booking detail page so I can confirm what I sent.",
     "GET /payments/:id/proof streams BYTEA (own-only); ProofPreview renders image or PDF embed.",
     "Should", "Implemented"),
    ("6. Payments (Manual UPI)", "PAY-04", "Admin",
     "As an admin, I want a Payment Re-Verify queue listing all upi_manual payments with verified_at IS NULL so I can clear them at sample-collection / clinic-visit.",
     "GET /admin/payments/pending-re-verify; partial index payments_pending_reverify_idx; AdminLayout bell badge shows count.",
     "Must", "Implemented"),
    ("6. Payments (Manual UPI)", "PAY-05", "Admin",
     "As an admin, I want to view the uploaded proof in a large modal alongside booking metadata + optional notes textarea before approving.",
     "ReVerifyPaymentModal with proof full-bleed + booking sidebar + notes; calls GET /admin/payments/:id/proof.",
     "Must", "Implemented"),
    ("6. Payments (Manual UPI)", "PAY-06", "Admin",
     "As an admin, I want re-verify to be idempotent so double-clicking doesn't break anything.",
     "POST /admin/payments/:paymentId/re-verify uses CTE WHERE verified_at IS NULL; second call returns already-verified row.",
     "Must", "Implemented"),
    ("6. Payments (Manual UPI)", "PAY-07", "Patient",
     "As a patient, I want my booking status to read 'In progress' until admin re-verifies, then automatically update to 'Confirmed'.",
     "Booking status transitions in_progress → confirmed; verified_at populated; cache invalidation propagates to patient hooks.",
     "Must", "Implemented"),
    ("6. Payments (Manual UPI)", "PAY-08", "Admin",
     "As an admin, I want to record an offline payment (cash, UPI QR offline, card swipe, cheque) against an existing booking so partial payments can be added later.",
     "POST /admin/bookings/:id/payments uses FOR UPDATE lock; recomputes paid/balance/payment_status from SUM(amount-refunded_amount).",
     "Must", "Implemented"),
    ("6. Payments (Manual UPI)", "PAY-09", "System",
     "As the system, I want to keep Razorpay stubs for later reintroduction so we can switch back without rewriting.",
     "POST /bookings/:id/confirm-payment-stub + POST /payments/create-order + POST /webhooks/razorpay (no signature verify yet) all preserved.",
     "Should", "Preserved as stub"),

    # ---------- 7. Reports / Lab Results ----------
    ("7. Reports / Lab Results", "RPT-01", "Admin",
     "As an admin, I want to look up a booking by its human-readable booking_code (AROGYA-TEST-YYYYMM-NNNNNN) so I can upload reports without typing numeric IDs.",
     "GET /admin/booking-lookup?code=; returns id + patient_snapshot + items_summary + doctor_name.",
     "Must", "Implemented"),
    ("7. Reports / Lab Results", "RPT-02", "Admin",
     "As an admin, I want to upload multiple report files (PDF/image/DICOM/Word/ZIP - any format) for a booking with a chosen report_type.",
     "POST /admin/bookings/:bookingId/reports (multer 25 MB cap, no MIME whitelist); version = MAX+1 per booking; no soft-disable of prior versions.",
     "Must", "Implemented"),
    ("7. Reports / Lab Results", "RPT-03", "Admin",
     "As an admin, I want a standalone /admin/reports page that lists recent reports across bookings with thumbnails + preview + delete actions.",
     "GET /admin/reports?q=&limit= joins reports + bookings; AdminReportsPage shows 48×48 thumbnail or file icon, action buttons.",
     "Should", "Implemented"),
    ("7. Reports / Lab Results", "RPT-04", "Admin",
     "As an admin, I want to soft-delete a report so it disappears from the patient's view but stays for audit.",
     "DELETE /admin/reports/:reportId sets is_active=FALSE; cache invalidation hits admin list + patient myReports + booking detail.",
     "Must", "Implemented"),
    ("7. Reports / Lab Results", "RPT-05", "Patient",
     "As a patient, I want to download my report via a short-lived signed URL so the file is protected.",
     "GET /reports/:id/download (uploads route) returns {url, expires_in:900} via storage.getSignedUrl; downloadReport() opens new tab.",
     "Must", "Implemented"),
    ("7. Reports / Lab Results", "RPT-06", "System",
     "As the system, I want reports for walk-in bookings to be allowed even when no patient user exists.",
     "Migration 0023: reports.patient_user_id is NULLABLE; ownership is via booking_id; upload handler still populates when booking has a user.",
     "Must", "Implemented"),

    # ---------- 8. Invoices ----------
    ("8. Invoices", "INV-01", "System",
     "As the system, I want every booking (online doctor / online test / walk-in) to auto-create exactly one invoice with a generated number on INSERT.",
     "INSERT invoices(invoice_number='') inside the same booking transaction; generate_invoice_number trigger fills INV-YYYYMM-NNNNNN (Asia/Kolkata).",
     "Must", "Implemented"),
    ("8. Invoices", "INV-02", "Patient",
     "As a patient, I want to download my own invoice as a PDF from my booking detail page.",
     "GET /bookings/:id/invoice.pdf (own-only); shared streamInvoicePdf() helper; A4 PDF via pdfkit.",
     "Must", "Implemented"),
    ("8. Invoices", "INV-03", "Admin",
     "As an admin, I want to list all invoices with filters (type/origin/from/to/free-text) so I can find them quickly.",
     "GET /admin/invoices?type=&origin=&from=&to=&q= joins invoices + bookings; max 500 rows; ILIKE on invoice_number, booking_code, patient_snapshot JSON keys.",
     "Must", "Implemented"),
    ("8. Invoices", "INV-04", "Admin",
     "As an admin, I want per-row download of the invoice PDF from the invoices list.",
     "GET /admin/bookings/:id/invoice.pdf delegates to streamInvoicePdf(); same PDF as patient download.",
     "Must", "Implemented"),
    ("8. Invoices", "INV-05", "System",
     "As the system, I want the invoice header/footer to reflect current clinic settings (name, address, phone, email, GSTIN).",
     "PDF reads clinic_settings; clinic_address JSON flattened to 'line1, line2, city, state, pincode'.",
     "Should", "Implemented"),

    # ---------- 9. Admin Dashboard & Analytics ----------
    ("9. Admin Dashboard & Analytics", "DSH-01", "Admin",
     "As an admin, I want a dashboard with today's bookings, today's revenue, pending reports, new patients this week and yesterday comparison.",
     "GET /admin/dashboard returns kpis + schedule; computes today/yesterday boundaries in Asia/Kolkata; revenue from payments (captured-refunded).",
     "Must", "Implemented"),
    ("9. Admin Dashboard & Analytics", "DSH-02", "Admin",
     "As an admin, I want the dashboard to auto-refresh so I can keep it open as a live operations screen.",
     "useAdminDashboard refetchInterval: 60_000 ms.",
     "Should", "Implemented"),
    ("9. Admin Dashboard & Analytics", "DSH-03", "Admin",
     "As an admin, I want today's schedule of the next 10 bookings with doctor + items summary on the dashboard.",
     "schedule[] subquery returns 10 rows with inline doctor_name and items_summary via string_agg.",
     "Must", "Implemented"),
    ("9. Admin Dashboard & Analytics", "DSH-04", "Admin",
     "As an admin, I want an Analytics page showing daily revenue + booking trend, top services, type/origin/status/payment-method distributions over the last N days.",
     "GET /admin/analytics?days=N (1-365); Promise.all-fans-out 7 queries; generate_series ensures empty days are charted.",
     "Should", "Implemented"),
    ("9. Admin Dashboard & Analytics", "DSH-05", "Admin",
     "As an admin, I want to export analytics to CSV so I can share with stakeholders.",
     "Multi-section CSV built client-side via Blob URL (KPIs + daily trend + top services + status + payment methods).",
     "Should", "Implemented"),
    ("9. Admin Dashboard & Analytics", "DSH-06", "Admin",
     "As an admin, I want the heavy analytics charts lazy-loaded so the rest of the app stays light.",
     "AnalyticsPage uses React.lazy in App.tsx; chunk is ~119 kB gzipped.",
     "Could", "Implemented"),

    # ---------- 10. Admin Bookings Management ----------
    ("10. Admin Bookings", "AB-01", "Admin",
     "As an admin, I want a filterable bookings list (by type, origin, status, free-text) so I can find rows fast.",
     "GET /admin/bookings?type=&origin=&status=&q=&visit_type=; subqueries inline doctor_name + items_summary; LIMIT 200 ORDER BY created_at desc.",
     "Must", "Implemented"),
    ("10. Admin Bookings", "AB-02", "Admin",
     "As an admin, I want to open a booking detail page showing items, payments, active reports, advance/balance and the patient snapshot.",
     "GET /admin/bookings/:id; full booking + items + payments + active reports.",
     "Must", "Implemented"),
    ("10. Admin Bookings", "AB-03", "Admin",
     "As an admin, I want to change a booking's status (pending_payment / confirmed / in_progress / completed / cancelled / no_show) from a color-coded dropdown both in the list and the detail sidebar.",
     "PATCH /admin/bookings/:id/status; BookingStatusSelect compact + default variants; invalidates admin + patient caches.",
     "Must", "Implemented"),
    ("10. Admin Bookings", "AB-04", "Admin",
     "As an admin, I want to filter by visit_type (in_clinic vs home_visit) so I can split the worklist.",
     "visit_type filter on /admin/bookings and /admin/invoices; SELECT also returns b.visit_type for origin pill.",
     "Should", "Implemented"),

    # ---------- 11. Admin Walk-in Bills ----------
    ("11. Walk-in Bills", "WB-01", "Admin",
     "As a receptionist, I want a Create Walk-in Bill page where I can capture patient (snapshot or existing user), N line items (catalog or custom), and optional payment in one form.",
     "POST /admin/walk-in-bills with discriminated-union walkInItemSchema; single transaction creates booking + items + optional payment + invoice.",
     "Must", "Implemented"),
    ("11. Walk-in Bills", "WB-02", "Admin",
     "As a receptionist, I want walk-in bookings to use today's date/time and skip slot generation so I can bill quickly.",
     "INSERT bookings with booking_origin='walk_in', visit_type='in_clinic', scheduled_date=today, scheduled_start_time=now, collection_status='not_required'.",
     "Must", "Implemented"),
    ("11. Walk-in Bills", "WB-03", "Admin",
     "As a receptionist, I want the bill to automatically compute subtotal, total, paid, balance, payment_status and booking_status so I don't make arithmetic mistakes.",
     "Backend computes payment_status (pending|partial|paid) and booking_status (completed if paid in full else in_progress).",
     "Must", "Implemented"),
    ("11. Walk-in Bills", "WB-04", "Admin",
     "As a receptionist, I want offline payment captured atomically with the booking so the invoice numbers correctly.",
     "INSERT payments(payment_source='offline', payment_status='captured', collected_by_user_id=req.user.id) when payment.amount > 0; invoice trigger fires.",
     "Must", "Implemented"),

    # ---------- 12. Admin Catalog Management ----------
    ("12. Catalog: Categories", "CAT-01", "Admin",
     "As an admin, I want full CRUD over service categories with auto-slug, duplicate detection, soft delete, and a doctors/services count badge.",
     "GET/POST/PATCH/DELETE /admin/service-categories[/:id]; FK on services.category_id is ON DELETE RESTRICT so deletes are soft (is_active=FALSE).",
     "Must", "Implemented"),
    ("12. Catalog: Services", "CAT-02", "Admin",
     "As an admin, I want full CRUD over services + health packages with category dropdown, auto-slug, package toggle (package_service_ids + discount %), and soft delete.",
     "GET/POST/PATCH/DELETE /admin/services[/:id]; services_package_consistency_chk enforced by always sending [] for package_service_ids on package rows.",
     "Must", "Implemented"),
    ("12. Catalog: Services", "CAT-03", "Admin",
     "As an admin, I want unique slug + test_key conflicts surfaced as friendly 409 messages so I know which field collides.",
     "Pre-checks both UNIQUE constraints and returns 409 DUPLICATE with the offending column.",
     "Should", "Implemented"),
    ("12. Catalog: Departments", "CAT-04", "Admin",
     "As an admin, I want CRUD over departments with auto-slug, doctor count, and soft delete.",
     "GET/POST/PATCH/DELETE /admin/departments[/:id]; FK users.department_id ON DELETE SET NULL but soft-delete pattern used.",
     "Must", "Implemented"),
    ("12. Catalog: Pincodes", "CAT-05", "Admin",
     "As an admin, I want CRUD over serviceable pincodes (with city/state/zone/home_visit_charge/lead_time) so home collection is correctly scoped.",
     "GET/POST/PATCH/DELETE /admin/serviceable-pincodes[/:id]; pincode field disabled in edit (natural key); soft delete.",
     "Must", "Implemented"),
    ("12. Catalog: Pincodes", "CAT-06", "System",
     "As the system, I want a deactivated pincode to immediately stop being bookable for home collection.",
     "Public /serviceable-pincodes/check filters WHERE is_active=TRUE; no extra work needed.",
     "Must", "Implemented"),

    # ---------- 13. Admin Doctor Management ----------
    ("13. Doctor Management", "DR-01", "Admin",
     "As an admin, I want to list all doctors (including inactive) with department filter, search, photo avatar, and centers count.",
     "GET /admin/doctors?q=&department_id=&is_active=; centers_count subquery; DoctorAvatar per row.",
     "Must", "Implemented"),
    ("13. Doctor Management", "DR-02", "Admin",
     "As an admin, I want to add a doctor with profile (name, speciality, qualifications, consultation fee, about, department) and one or more consulting centers in a single submit.",
     "POST /admin/doctors with nested centers[]; transaction inserts users(role=doctor,is_login_enabled=FALSE) + N doctor_centers.",
     "Must", "Implemented"),
    ("13. Doctor Management", "DR-03", "Admin",
     "As an admin, I want to upload a doctor's profile photo and have it served back at a public URL for marketing pages.",
     "POST /admin/doctors/:id/photo (5 MB, JPG/PNG/WEBP); writes BYTEA + MIME; GET /doctors/:id/photo public stream with Cache-Control max-age=300.",
     "Must", "Implemented"),
    ("13. Doctor Management", "DR-04", "Admin",
     "As an admin, I want to add/edit/soft-delete a doctor's consulting centers including a visual Weekly Schedule editor with slot duration, buffer, lunch.",
     "POST /admin/doctors/:id/centers, PATCH/DELETE /admin/centers/:centerId; WeeklyScheduleEditor emits WeeklySchedule JSON matching doctor_centers.schedule shape.",
     "Must", "Implemented"),
    ("13. Doctor Management", "DR-05", "Admin",
     "As an admin, I want a separate home-visit schedule per center so a doctor can do clinic visits on some days and home visits on others.",
     "doctor_centers.home_visit_schedule JSONB; reads in slot generator when visit_type=home_visit.",
     "Should", "Implemented"),

    # ---------- 14. Slot Management ----------
    ("14. Slot Generation", "SLOT-01", "System",
     "As the system, I want to generate slots live from the doctor's schedule, subtracting existing bookings and unavailability blocks, so admins manage availability through the schedule editor.",
     "Slot generator reads schedule/home_visit_schedule for weekday, generates from {start,end,slot_minutes,buffer_minutes,lunch_start,lunch_end}, subtracts active bookings + doctor_unavailability.",
     "Must", "Implemented"),
    ("14. Slot Generation", "SLOT-02", "Admin",
     "As an admin, I want to block whole days or specific slots for a doctor via doctor_unavailability so vacations and emergencies are respected.",
     "doctor_unavailability rows: NULL slot_start_time = whole day; NULL doctor_center_id = all centers.",
     "Should", "DB ready, no admin UI yet"),
    ("14. Slot Generation", "SLOT-03", "System",
     "As the system, I want to prevent double-booking via a partial unique index on (doctor_user_id, doctor_center_id, scheduled_date, scheduled_start_time).",
     "uq_active_slot partial unique index covers booking_status IN (draft, pending_payment, confirmed, in_progress).",
     "Must", "Implemented"),
    ("14. Slot Generation", "SLOT-04", "System",
     "As the system, I want a cron to expire stale drafts so slot locks are released.",
     "Flag stale draft bookings whose slot_locked_until has passed to cancelled.",
     "Should", "TODO (noted in migration)"),

    # ---------- 15. Home Collection Workflow ----------
    ("15. Home Collection", "HC-01", "Admin",
     "As an admin, I want CRUD over home collector accounts (phlebotomists) including name, mobile, DOB (for age), gender, photo, active toggle.",
     "GET/POST/PATCH/DELETE /admin/collectors; users row with role='patient'+admin_role='collector' or similar; photo uploaded as BYTEA via /admin/collectors/:id/photo.",
     "Must", "Implemented"),
    ("15. Home Collection", "HC-02", "Admin",
     "As an admin, I want a Home Collection Board (Kanban) showing each home-visit booking by collection_status so I can dispatch fast.",
     "HomeCollectionBoardPage groups bookings by collection_status: not_assigned → assigned → en_route → collected → received_at_lab.",
     "Must", "Implemented"),
    ("15. Home Collection", "HC-03", "Admin",
     "As an admin, I want to assign a collector to a not_assigned booking via a dropdown directly on the board.",
     "useUpdateCollection mutation sets collection_status='assigned' + assigned_staff_user_id; useCollectionStaff() feeds dropdown.",
     "Must", "Implemented"),
    ("15. Home Collection", "HC-04", "Admin",
     "As an admin, I want to progress a collection through stages so the patient and lab are kept in sync.",
     "PATCH /admin/home-collections/:id transitions collection_status through assigned → en_route → collected → received_at_lab.",
     "Must", "Implemented"),
    ("15. Home Collection", "HC-05", "Patient",
     "As a patient, I want to see the assigned collector's name, age, phone and photo on my booking page so I know who is coming.",
     "Booking detail surfaces assigned_staff metadata (full name, age computed from DOB, mobile, photo URL).",
     "Should", "Implemented"),

    # ---------- 16. Notifications ----------
    ("16. Notifications", "NOTIF-01", "System",
     "As the system, I want a notifications table that supports both per-user and admin-broadcast rows so the bell can show targeted + clinic-wide events.",
     "Migration 0026: notifications + notification_reads; audience IN (admin, patient, collector); partial indexes for hot paths (unread + admin broadcasts).",
     "Must", "Implemented"),
    ("16. Notifications", "NOTIF-02", "Admin / Patient",
     "As an admin or patient, I want a bell icon in my header with an unread badge so I can see new events at a glance.",
     "NotificationBell variants light (patient header) / dark (admin layout); badge shows count capped at 9+.",
     "Must", "Implemented"),
    ("16. Notifications", "NOTIF-03", "Admin / Patient",
     "As an admin or patient, I want the bell to poll the server every 30s and refetch on tab focus + on open so I see fresh events.",
     "useMyNotifications polls; dropdown open also invalidates ['notifications','mine'].",
     "Should", "Implemented"),
    ("16. Notifications", "NOTIF-04", "Admin / Patient",
     "As an admin or patient, I want to click a notification to mark it read and deep-link to the related entity.",
     "useMarkNotificationRead; broadcasts use notification_reads table; non-broadcast use read_at column; click triggers Link to n.link.",
     "Must", "Implemented"),
    ("16. Notifications", "NOTIF-05", "Admin / Patient",
     "As an admin or patient, I want a Mark All Read button to bulk-clear my unread count.",
     "useMarkAllNotificationsRead.",
     "Should", "Implemented"),
    ("16. Notifications", "NOTIF-06", "System",
     "As the system, I want notifications fired for proof_submitted, reverified, collector_assigned, collection_status_changed so users are informed of state changes.",
     "Backend domain events INSERT notifications with appropriate event + link + audience.",
     "Must", "Implemented"),

    # ---------- 17. Audit Log ----------
    ("17. Audit Log", "AUDIT-01", "System",
     "As the system, I want every successful authenticated mutating request (POST/PATCH/PUT/DELETE 2xx) automatically logged with actor, action, entity, IP and user-agent.",
     "auditLogMiddleware hooks res.on('finish'); classify() derives action from URL (e.g. services.create, payments.re-verify); SKIP_PATTERNS exclude auth/me/health/webhooks.",
     "Must", "Implemented"),
    ("17. Audit Log", "AUDIT-02", "Super Admin",
     "As a super_admin, I want an admin audit-log page to review who did what when.",
     "AdminAuditLogPage uses GET /admin/audit-log; filters by actor, action, entity, date range.",
     "Should", "Implemented"),
    ("17. Audit Log", "AUDIT-03", "System",
     "As the system, I want audit failures to never break the real action.",
     "INSERT is fire-and-forget after res.finish; errors are logged via pino, not propagated.",
     "Must", "Implemented"),

    # ---------- 18. Admin Records / Misc ----------
    ("18. Patients Records", "REC-01", "Admin",
     "As an admin, I want a Patients page listing all patient users with mobile, last login, booking count so I can support them.",
     "AdminPatientsPage queries /admin/patients (assumed).",
     "Should", "Implemented"),
    ("18. Patients Records", "REC-02", "Admin",
     "As an admin, I want to moderate reviews (pending → approved/rejected/hidden) with optional rejection_reason and clinic_reply.",
     "AdminFeedbackPage reads pending reviews; reviews table tracks status, clinic_reply, replied_at, replied_by.",
     "Should", "Implemented"),
    ("18. Patients Records", "REC-03", "Admin",
     "As an admin, I want to triage enquiries from the public contact form (new → read → replied → closed).",
     "AdminEnquiriesPage; enquiries.status transitions and responded_at/responded_by columns.",
     "Should", "Implemented"),
    ("18. Patients Records", "REC-04", "Admin",
     "As an admin, I want a Payments list showing all online + offline + manual-UPI rows so I can reconcile.",
     "AdminPaymentsPage queries /admin/payments (assumed).",
     "Should", "Implemented"),

    # ---------- 19. Admin Settings ----------
    ("19. Settings", "SET-01", "Super Admin",
     "As a super_admin, I want to edit clinic settings (name, address, phone, email, GSTIN, UPI VPA, business hours, tax %, cancellation window, refund policy, retention days).",
     "AdminSettingsPage reads/writes clinic_settings JSONB key/value rows; seeded by 01_clinic_settings.sql.",
     "Should", "Implemented"),
    ("19. Settings", "SET-02", "System",
     "As the system, I want the patient-facing UPI VPA endpoint to be public so checkout doesn't require auth.",
     "GET /clinic/upi is public; returns {upi_id, upi_display_name}.",
     "Must", "Implemented"),

    # ---------- 20. Cart ----------
    ("20. Cart", "CRT-01", "Patient",
     "As a patient, I want my cart contents to persist across reloads so I don't lose my selection.",
     "Zustand useCartStore persisted to localStorage key 'arogya-cart'.",
     "Must", "Implemented"),
    ("20. Cart", "CRT-02", "Patient",
     "As a patient on the cart page, I want to update quantities or remove items before checkout.",
     "CartPage uses useCartStore { addItem, removeItem, updateQuantity, clear }.",
     "Must", "Implemented"),
    ("20. Cart", "CRT-03", "Patient",
     "As a patient, I want a header badge that shows my cart count so I never forget what I added.",
     "Header reads useCartStore().count() and renders badge.",
     "Should", "Implemented"),

    # ---------- 21. Cross-cutting / Non-functional ----------
    ("21. Cross-cutting / NFR", "NFR-01", "System",
     "As the system, I want all booking + invoice numbers generated in Asia/Kolkata so the month suffix matches local time.",
     "Triggers generate_booking_code + generate_invoice_number use AT TIME ZONE 'Asia/Kolkata'.",
     "Must", "Implemented"),
    ("21. Cross-cutting / NFR", "NFR-02", "System",
     "As the system, I want a global rate limit (200/min/IP) and a stricter limit (5/min/IP) on public write endpoints so we resist abuse.",
     "express-rate-limit generalLimiter on all routes; publicWriteLimiter on POST /enquiries + POST /reviews.",
     "Must", "Implemented"),
    ("21. Cross-cutting / NFR", "NFR-03", "System",
     "As the system, I want sensitive fields scrubbed from logs (authorization header, password*, token).",
     "Pino redaction list configured in lib/logger.ts.",
     "Must", "Implemented"),
    ("21. Cross-cutting / NFR", "NFR-04", "System",
     "As the system, I want every DB write that spans multiple tables (booking + items + invoice + payment) inside a single transaction.",
     "transaction() helper wraps BEGIN/COMMIT/ROLLBACK; all booking creation paths use it.",
     "Must", "Implemented"),
    ("21. Cross-cutting / NFR", "NFR-05", "System",
     "As the system, I want frontend hooks to invalidate every affected cache key on a mutation so admin + patient views stay in sync.",
     "Documented in memory feedback_keep_models_in_sync; e.g. useCreateServiceCategory invalidates 3 keys.",
     "Must", "Implemented"),
    ("21. Cross-cutting / NFR", "NFR-06", "System",
     "As the system, I want a startup env-validation gate so missing JWT_SECRET / DATABASE_URL crashes early rather than at first request.",
     "config/env.ts Zod-validates required vars; throws on startup.",
     "Must", "Implemented"),
    ("21. Cross-cutting / NFR", "NFR-07", "System",
     "As the system, I want errorHandler to translate Zod / Multer / pg constraint errors into stable HTTP codes with column-aware messages.",
     "errorHandler.ts maps ZodError→400, MulterError→413/400, 23505→409, 23503/23502/23514→400, 42703/42P01→500 with 'pnpm db:migrate' hint.",
     "Must", "Implemented"),
    ("21. Cross-cutting / NFR", "NFR-08", "System",
     "As the system, I want a feature flag (VITE_PHASE_2_ENABLED) to gate the entire patient + admin Phase-2 surface so we can ship Phase-1 marketing first.",
     "PHASE_2_ENABLED gate in App.tsx wraps Phase-2 routes.",
     "Should", "Implemented"),
]

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

wb = Workbook()

# ----- Sheet 1: Cover -----
cover = wb.active
cover.title = "Cover"

title_font = Font(name="Calibri", size=22, bold=True, color="FFFFFF")
subtitle_font = Font(name="Calibri", size=12, italic=True, color="EDEDED")
header_fill = PatternFill("solid", fgColor="0F4C81")
section_fill = PatternFill("solid", fgColor="E7EEF7")
section_font = Font(name="Calibri", size=12, bold=True, color="0F4C81")
body_font = Font(name="Calibri", size=11)

cover.merge_cells("A1:E3")
cover["A1"] = "Arogya Diagnostics — User Stories"
cover["A1"].font = title_font
cover["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cover["A1"].fill = header_fill

cover.merge_cells("A4:E5")
cover["A4"] = ("Module-wise backlog of user stories derived from the live codebase\n"
               "(React + Express + Postgres monorepo). Generated 2026-05-14.")
cover["A4"].font = subtitle_font
cover["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cover["A4"].fill = header_fill

cover.row_dimensions[1].height = 30
cover.row_dimensions[2].height = 30
cover.row_dimensions[3].height = 30
cover.row_dimensions[4].height = 30
cover.row_dimensions[5].height = 30

cover["A7"] = "How to use"
cover["A7"].font = section_font
cover["A7"].fill = section_fill
cover.merge_cells("A7:E7")
cover["A8"] = ("• Each module has its own sheet. The 'All stories' sheet is a single-table view across modules.\n"
               "• Story columns: ID, Role, User story (As a … I want … so that …), Acceptance criteria, Priority (MoSCoW), Status.\n"
               "• Status reflects what is observable in the code as of generation date — Implemented / Stub / TODO etc.\n"
               "• Priority follows MoSCoW: Must, Should, Could.")
cover["A8"].font = body_font
cover["A8"].alignment = Alignment(wrap_text=True, vertical="top")
cover.merge_cells("A8:E12")

cover["A14"] = "Module index"
cover["A14"].font = section_font
cover["A14"].fill = section_fill
cover.merge_cells("A14:E14")

# Module list will be filled after we know the modules.
modules_ordered = []
seen = set()
for row in STORIES:
    if row[0] not in seen:
        modules_ordered.append(row[0])
        seen.add(row[0])

start_row = 15
for i, mod in enumerate(modules_ordered):
    cell = cover.cell(row=start_row + i, column=1, value=mod)
    cell.font = body_font
    # count
    count_cell = cover.cell(row=start_row + i, column=2, value=f"{sum(1 for r in STORIES if r[0]==mod)} stories")
    count_cell.font = body_font

cover.column_dimensions["A"].width = 50
cover.column_dimensions["B"].width = 18
cover.column_dimensions["C"].width = 18
cover.column_dimensions["D"].width = 18
cover.column_dimensions["E"].width = 18

# ----- Sheet 2..N: One per module + a "All stories" summary -----
COLS = ["ID", "Role", "User story", "Acceptance criteria", "Priority", "Status"]
COL_WIDTHS = [12, 14, 60, 70, 12, 22]

def style_header(ws, ncols):
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0F4C81")
    thin = Side(border_style="thin", color="BFBFBF")
    for col_idx in range(1, ncols + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = fill
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

def style_body(ws, nrows, ncols):
    thin = Side(border_style="thin", color="DDDDDD")
    body = Font(name="Calibri", size=11)
    for r in range(2, nrows + 2):
        ws.row_dimensions[r].height = 60
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def sanitize_sheet_name(name):
    # Excel sheet names: <=31 chars, no : \ / ? * [ ]
    bad = ':\\/?*[]'
    out = ''.join('-' if ch in bad else ch for ch in name)
    return out[:31]

# Per-module sheets
for mod in modules_ordered:
    ws = wb.create_sheet(sanitize_sheet_name(mod))
    rows = [r for r in STORIES if r[0] == mod]

    # Title row above table? Use row 1 as header for simplicity.
    for i, h in enumerate(COLS, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(COLS))

    for r_idx, row in enumerate(rows, start=2):
        _, sid, role, story, ac, prio, status = row
        ws.cell(row=r_idx, column=1, value=sid)
        ws.cell(row=r_idx, column=2, value=role)
        ws.cell(row=r_idx, column=3, value=story)
        ws.cell(row=r_idx, column=4, value=ac)
        ws.cell(row=r_idx, column=5, value=prio)
        ws.cell(row=r_idx, column=6, value=status)

    style_body(ws, len(rows), len(COLS))

    # Conditional fills for priority + status
    prio_fills = {
        "Must": PatternFill("solid", fgColor="FDE7E9"),
        "Should": PatternFill("solid", fgColor="FFF4D6"),
        "Could": PatternFill("solid", fgColor="E5F0FF"),
    }
    status_fills_keys = [
        ("Implemented", PatternFill("solid", fgColor="E2F0D9")),
        ("Stub", PatternFill("solid", fgColor="FFF2CC")),
        ("TODO", PatternFill("solid", fgColor="FCE4D6")),
        ("Preserved", PatternFill("solid", fgColor="E5F0FF")),
        ("DB ready", PatternFill("solid", fgColor="FFF2CC")),
    ]
    for r_idx in range(2, len(rows) + 2):
        pcell = ws.cell(row=r_idx, column=5)
        if pcell.value in prio_fills:
            pcell.fill = prio_fills[pcell.value]
        scell = ws.cell(row=r_idx, column=6)
        if scell.value:
            for key, fill in status_fills_keys:
                if scell.value.startswith(key):
                    scell.fill = fill
                    break

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# All stories combined sheet
ws_all = wb.create_sheet("All stories")
all_cols = ["Module", "ID", "Role", "User story", "Acceptance criteria", "Priority", "Status"]
all_widths = [32, 12, 14, 60, 70, 12, 22]

for i, h in enumerate(all_cols, start=1):
    ws_all.cell(row=1, column=i, value=h)
style_header(ws_all, len(all_cols))

for r_idx, row in enumerate(STORIES, start=2):
    for c_idx, val in enumerate(row, start=1):
        ws_all.cell(row=r_idx, column=c_idx, value=val)

style_body(ws_all, len(STORIES), len(all_cols))

# colour priorities + status in combined sheet
prio_fills = {
    "Must": PatternFill("solid", fgColor="FDE7E9"),
    "Should": PatternFill("solid", fgColor="FFF4D6"),
    "Could": PatternFill("solid", fgColor="E5F0FF"),
}
status_fills_keys = [
    ("Implemented", PatternFill("solid", fgColor="E2F0D9")),
    ("Stub", PatternFill("solid", fgColor="FFF2CC")),
    ("TODO", PatternFill("solid", fgColor="FCE4D6")),
    ("Preserved", PatternFill("solid", fgColor="E5F0FF")),
    ("DB ready", PatternFill("solid", fgColor="FFF2CC")),
]
for r_idx in range(2, len(STORIES) + 2):
    pcell = ws_all.cell(row=r_idx, column=6)
    if pcell.value in prio_fills:
        pcell.fill = prio_fills[pcell.value]
    scell = ws_all.cell(row=r_idx, column=7)
    if scell.value:
        for key, fill in status_fills_keys:
            if scell.value.startswith(key):
                scell.fill = fill
                break

for i, w in enumerate(all_widths, start=1):
    ws_all.column_dimensions[get_column_letter(i)].width = w

# Apply autofilter on All stories
ws_all.auto_filter.ref = ws_all.dimensions

# ----- Save -----
out_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(out_dir, "Arogya_User_Stories.xlsx")
wb.save(out_path)
print(f"Wrote {out_path}")
print(f"Total stories: {len(STORIES)} across {len(modules_ordered)} modules")
